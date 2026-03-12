#!/usr/bin/env python3
"""
STEP 1: PREPARE PROJECTIONS SPREADSHEET
========================================
Takes raw SC JSON + custom projection CSVs → produces master spreadsheet.

Input:
  - AFL_supercoach_playerlist_2026_R0  (JSON from SC website)
  - rookie_projections.csv             (custom rookie avg overrides)
  - midprice_projections.csv           (custom mid-price avg overrides)

Output:
  - SC2026_AllPlayers_Projections.xlsx  (master spreadsheet with User columns)
  - excluded_players.csv               (auto-generated from injury data in JSON)

The spreadsheet includes:
  - "Our Proj" = blended projection with all custom overrides baked in
  - "User Proj" = blank column for user to add final manual overrides
  - "User Incl/Excl" = blank column: 1 = MUST HAVE, 0 = MUST AVOID
"""
import json, csv, sys
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, Protection
from openpyxl.utils import get_column_letter

# =============================================================================
# CONFIGURATION
# =============================================================================
JSON_FILE = 'AFL_supercoach_playerlist_2026_R0'
ROOKIE_CSV = 'rookie_projections.csv'
MIDPRICE_CSV = 'midprice_projections.csv'
OUTPUT_XLSX = 'SC2026_AllPlayers_Projections.xlsx'
OUTPUT_EXCLUDED = 'excluded_players.csv'

TEAM_BYES = {
    'ADE': 0, 'BRL': 2, 'CAR': 2, 'COL': 2, 'ESS': 0, 'FRE': 0,
    'GCS': 3, 'GEE': 2, 'GWS': 4, 'HAW': 3, 'MEL': 0, 'NTH': 0,
    'PTA': 0, 'RIC': 0, 'STK': 4, 'SYD': 3, 'WBD': 3, 'WCE': 0,
}
OPENING_ROUND_CLUBS = {'SYD','CAR','GCS','GEE','GWS','HAW','BRL','WBD','STK','COL'}
PRICE_CHANGE_PER_PT = 440
GAMES_WEIGHT_CAP = 0.80

# =============================================================================
# 1. LOAD RAW DATA
# =============================================================================
print(f"Loading {JSON_FILE}...")
with open(JSON_FILE) as f:
    raw = json.load(f)
print(f"  {len(raw)} players loaded")

# =============================================================================
# 2. LOAD CUSTOM CSV OVERRIDES
# =============================================================================
rookie_overrides = {}
try:
    with open(ROOKIE_CSV) as f:
        for row in csv.DictReader(f):
            if row.get('exclude') == '1': continue
            try:
                avg = float(row['custom_avg'])
                if avg > 0: rookie_overrides[row['name']] = avg
            except: pass
    print(f"  {len(rookie_overrides)} rookie overrides from {ROOKIE_CSV}")
except FileNotFoundError:
    print(f"  {ROOKIE_CSV} not found, skipping")

midprice_overrides = {}
try:
    with open(MIDPRICE_CSV) as f:
        for row in csv.DictReader(f):
            try:
                avg = float(row['custom_avg'])
                if avg > 0: midprice_overrides[row['name']] = avg
            except: pass
    print(f"  {len(midprice_overrides)} midprice overrides from {MIDPRICE_CSV}")
except FileNotFoundError:
    print(f"  {MIDPRICE_CSV} not found, skipping")

# Build combined overrides from CSV files only (no hardcoded values)
custom_overrides = {**rookie_overrides, **midprice_overrides}
print(f"  Total custom overrides: {len(custom_overrides)}")

# =============================================================================
# 3. EXTRACT INJURY EXCLUSIONS FROM JSON
# =============================================================================
excluded = []
for p in raw:
    if p.get('injury_suspension_status'):
        name = f"{p['first_name']} {p['last_name']}"
        stats = p['player_stats'][0] if p['player_stats'] else {}
        excluded.append({
            'name': name,
            'team': p['team']['abbrev'],
            'price': stats.get('price', 0) if stats else 0,
            'own_pct': stats.get('own', 0) or 0 if stats else 0,
            'injury': p.get('injury_suspension_status', ''),
            'injury_text': p.get('injury_suspension_status_text', ''),
        })

with open(OUTPUT_EXCLUDED, 'w', newline='') as f:
    if excluded:
        w = csv.DictWriter(f, fieldnames=excluded[0].keys())
        w.writeheader()
        w.writerows(excluded)
print(f"  {len(excluded)} injury exclusions written to {OUTPUT_EXCLUDED}")
excluded_names = {e['name'] for e in excluded}

# =============================================================================
# 4. BUILD PLAYER DATABASE WITH BLENDED PROJECTIONS
# =============================================================================
players = []
for p in raw:
    stats = p['player_stats'][0] if p['player_stats'] else {}
    if not stats or stats.get('price', 0) == 0: continue

    name = f"{p['first_name']} {p['last_name']}"
    team = p['team']['abbrev']
    positions = [pos['position'] for pos in p.get('positions', [])]
    price = stats['price']
    prev_avg = p.get('previous_average', 0) or 0
    prev_games = p.get('previous_games', 0) or 0
    own = stats.get('own', 0) or 0

    # SC Plus per-game projection (exclude bye-round zeros)
    ppts1 = stats.get('ppts1', 0) or 0
    ppts2 = stats.get('ppts2', 0) or 0
    ppts3 = stats.get('ppts3', 0) or 0
    playing_ppts = [x for x in [ppts1, ppts2, ppts3] if x > 0]
    sc_per_game = sum(playing_ppts) / len(playing_ppts) if playing_ppts else 0

    # Blend
    if name in custom_overrides:
        proj = custom_overrides[name]
        source = 'CUSTOM'
    else:
        gw = min(prev_games / 22.0, 1.0) * GAMES_WEIGHT_CAP
        if prev_avg > 0 and sc_per_game > 0:
            proj = gw * prev_avg + (1 - gw) * sc_per_game
        elif prev_avg > 0:
            proj = prev_avg
        elif sc_per_game > 0:
            proj = sc_per_game
        else:
            proj = 0
        source = 'BLENDED'

    bye = TEAM_BYES.get(team, 0)
    sgm = 22.5 if team in OPENING_ROUND_CLUBS else 23.0
    be = price / 5400
    games_3r = 2 if bye in (1, 2, 3, 4) else 3
    proj_rise = int((proj - be) * PRICE_CHANGE_PER_PT * games_3r)

    r0_score = stats.get('points', 0) or 0
    r0_games = stats.get('games', 0) or 0
    r0_vs_proj = r0_score - proj if r0_games > 0 else None

    # Flags
    flags = []
    if r0_games > 0 and r0_score > proj * 1.15: flags.append('HOT_R0')
    if prev_games > 0 and prev_games < 15 and prev_avg > be + 10: flags.append('INJURY_DISCOUNT')
    if name in excluded_names: flags.append('EXCLUDED')

    players.append({
        'name': name, 'team': team, 'pos': '/'.join(positions),
        'price': price, 'own': own, 'prev_avg': prev_avg, 'prev_games': prev_games,
        'sc_per_game': sc_per_game, 'proj': proj, 'source': source,
        'be': be, 'sgm': sgm, 'bye': bye, 'games_3r': games_3r,
        'proj_rise': proj_rise,
        'r0_score': r0_score if r0_games > 0 else None,
        'r0_vs_proj': r0_vs_proj,
        'flags': ', '.join(flags) if flags else '',
    })

players.sort(key=lambda r: -r['proj'] * r['sgm'])
print(f"  {len(players)} players with projections")

# =============================================================================
# 5. BUILD SPREADSHEET
# =============================================================================
wb = Workbook()
ws = wb.active
ws.title = "All Players"

hdr_font = Font(bold=True, color='FFFFFF', size=9, name='Arial')
hdr_fill = PatternFill('solid', fgColor='2F5496')
custom_fill = PatternFill('solid', fgColor='FFF2CC')
hot_fill = PatternFill('solid', fgColor='FCE4EC')
user_fill = PatternFill('solid', fgColor='E8F5E9')  # light green for user columns
excl_font = Font(size=9, name='Arial', color='AAAAAA')
thin = Border(left=Side('thin', 'D9D9D9'), right=Side('thin', 'D9D9D9'),
              top=Side('thin', 'D9D9D9'), bottom=Side('thin', 'D9D9D9'))

headers = [
    'Name', 'Team', 'Pos', 'Price', 'Own%', '2025 Avg', '2025 Gm',
    'SC PerGm', 'Our Proj', 'Source', 'BE', 'SC Games', 'Bye',
    'R0 Score', 'R0 vs Proj', 'Proj Rise', 'Flags',
    'User Proj', 'User Incl/Excl',
]
widths = [26, 5, 10, 11, 6, 8, 7, 8, 8, 7, 7, 7, 4, 8, 9, 10, 25, 10, 12]

for j, h in enumerate(headers):
    c = ws.cell(row=1, column=j+1, value=h)
    c.font = hdr_font
    c.fill = hdr_fill if j < 17 else PatternFill('solid', fgColor='2E7D32')
    c.alignment = Alignment(horizontal='center')
    c.border = thin
    ws.column_dimensions[get_column_letter(j+1)].width = widths[j]

for i, r in enumerate(players):
    dr = i + 2
    vals = [
        r['name'], r['team'], r['pos'], r['price'], r['own'],
        r['prev_avg'] or '', r['prev_games'] or '',
        round(r['sc_per_game'], 1) if r['sc_per_game'] else '',
        round(r['proj'], 1), r['source'], round(r['be'], 1),
        r['sgm'], f"R{r['bye']}" if r['bye'] > 0 else '-',
        r['r0_score'] if r['r0_score'] is not None else '',
        round(r['r0_vs_proj'], 1) if r['r0_vs_proj'] is not None else '',
        r['proj_rise'], r['flags'],
        '',  # User Proj (blank)
        '',  # User Incl/Excl (blank)
    ]
    for j, v in enumerate(vals):
        c = ws.cell(row=dr, column=j+1, value=v)
        c.font = Font(size=9, name='Arial')
        c.border = thin

    ws.cell(row=dr, column=4).number_format = '$#,##0'
    ws.cell(row=dr, column=5).number_format = '0.0'
    ws.cell(row=dr, column=16).number_format = '$#,##0'

    # Highlight user columns
    ws.cell(row=dr, column=18).fill = user_fill
    ws.cell(row=dr, column=19).fill = user_fill

    # Style flags
    if 'EXCLUDED' in r['flags']:
        for j in range(17):
            ws.cell(row=dr, column=j+1).font = excl_font
    elif r['source'] == 'CUSTOM':
        ws.cell(row=dr, column=9).fill = custom_fill
    if r['r0_vs_proj'] and r['r0_vs_proj'] > 15:
        ws.cell(row=dr, column=14).fill = hot_fill
        ws.cell(row=dr, column=15).fill = hot_fill

ws.auto_filter.ref = f"A1:S{len(players)+1}"
ws.freeze_panes = 'A2'

# Add instructions sheet
ws2 = wb.create_sheet("Instructions")
instructions = [
    "HOW TO USE THIS SPREADSHEET",
    "",
    "This spreadsheet is the input for the optimiser (run_optimiser.py).",
    "",
    "COLUMNS:",
    "  'Our Proj' = Blended projection with all custom overrides baked in.",
    "  'User Proj' = YOUR manual override. Leave blank to use 'Our Proj'.",
    "               The optimiser will use 'User Proj' if filled, otherwise 'Our Proj'.",
    "  'User Incl/Excl' = Hard constraints for the optimiser:",
    "               1 = MUST HAVE (force this player into the team)",
    "               0 = MUST AVOID (exclude this player completely)",
    "               blank = let the optimiser decide",
    "",
    "WORKFLOW:",
    "  1. Review projections in 'Our Proj' column",
    "  2. Add any overrides in 'User Proj' (green column)",
    "  3. Mark must-haves with 1 and must-avoids with 0 in 'User Incl/Excl'",
    "  4. Save the spreadsheet",
    "  5. Run: python run_optimiser.py",
    "",
    "NOTES:",
    "  - Players flagged as EXCLUDED (injured) are greyed out but not auto-excluded.",
    "    Mark them with 0 in User Incl/Excl if you want to exclude them.",
    "  - Players not named in R1 teams should also be marked with 0.",
    "  - Source 'CUSTOM' = override from rookie/midprice/ruck analysis.",
    "    Source 'BLENDED' = automatic blend of season avg + SC Plus per-game.",
    "  - HOT_R0 flag = scored >15% above projection in Opening Round.",
    "  - INJURY_DISCOUNT flag = low 2025 games but high avg, priced below ability.",
]
for i, line in enumerate(instructions):
    ws2.cell(row=i+1, column=1, value=line).font = Font(
        size=10 if i == 0 else 9,
        bold=(i == 0),
        name='Arial'
    )
ws2.column_dimensions['A'].width = 80

wb.save(OUTPUT_XLSX)
print(f"\nOutput: {OUTPUT_XLSX} ({len(players)} players)")
print(f"Output: {OUTPUT_EXCLUDED} ({len(excluded)} injuries)")
print(f"\nNext step: Review the spreadsheet, add User Proj / User Incl/Excl, then run run_optimiser.py")

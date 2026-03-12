#!/usr/bin/env python3
"""
STEP 2: RUN OPTIMISER
=====================
Reads projections from SC2026_AllPlayers_Projections.xlsx,
applies user overrides, runs dual-objective optimizer + Monte Carlo.

Input:
  - SC2026_AllPlayers_Projections.xlsx (from prepare_projections.py, with user edits)

The optimiser reads:
  - "Our Proj" as the base projection
  - "User Proj" as override (used instead of Our Proj if filled)
  - "User Incl/Excl": 1 = MUST HAVE, 0 = MUST AVOID, blank = optimiser decides

Output:
  - Console: optimal teams, Monte Carlo confidence, robustness analysis
"""
import random
from collections import defaultdict, Counter
from openpyxl import load_workbook
from pulp import *
try:
    from tqdm import tqdm
except ImportError:
    tqdm = None

# =============================================================================
# CONFIGURATION
# =============================================================================
INPUT_XLSX = 'SC2026_AllPlayers_Projections.xlsx'

OPENING_ROUND_CLUBS = {'SYD','CAR','GCS','GEE','GWS','HAW','BRL','WBD','STK','COL'}
PRICE_CHANGE_PER_PT = 440
TEAM_BYES = {
    'ADE': 0, 'BRL': 2, 'CAR': 2, 'COL': 2, 'ESS': 0, 'FRE': 0,
    'GCS': 3, 'GEE': 2, 'GWS': 4, 'HAW': 3, 'MEL': 0, 'NTH': 0,
    'PTA': 0, 'RIC': 0, 'STK': 4, 'SYD': 3, 'WBD': 3, 'WCE': 0,
}

# =============================================================================
# 1. LOAD SPREADSHEET
# =============================================================================
print(f"Loading {INPUT_XLSX}...")
wb = load_workbook(INPUT_XLSX, data_only=True)
ws = wb['All Players']

# Parse headers
headers = {}
for j in range(1, ws.max_column + 1):
    val = ws.cell(row=1, column=j).value
    if val:
        headers[val] = j

required = ['Name', 'Team', 'Pos', 'Price', 'Own%', 'Our Proj', 'Source',
            'SC Games', 'Bye', 'R0 Score', 'Proj Rise',
            'User Proj', 'User Incl/Excl']
for h in required:
    if h not in headers:
        print(f"ERROR: Missing column '{h}' in spreadsheet")
        exit(1)

# Parse players
players = []
must_have = []
must_avoid = []

for row in range(2, ws.max_row + 1):
    name = ws.cell(row=row, column=headers['Name']).value
    if not name: continue

    team = ws.cell(row=row, column=headers['Team']).value or ''
    pos_str = ws.cell(row=row, column=headers['Pos']).value or ''
    positions = [p.strip() for p in pos_str.split('/') if p.strip()]

    try:
        price = float(ws.cell(row=row, column=headers['Price']).value or 0)
    except (ValueError, TypeError):
        continue
    if price == 0: continue

    own = float(ws.cell(row=row, column=headers['Own%']).value or 0)
    our_proj = float(ws.cell(row=row, column=headers['Our Proj']).value or 0)
    source = ws.cell(row=row, column=headers['Source']).value or 'BLENDED'

    # User overrides
    user_proj_raw = ws.cell(row=row, column=headers['User Proj']).value
    user_incl_raw = ws.cell(row=row, column=headers['User Incl/Excl']).value

    # Apply user projection override
    if user_proj_raw is not None and user_proj_raw != '':
        try:
            proj = float(user_proj_raw)
            source = 'USER'
        except (ValueError, TypeError):
            proj = our_proj
    else:
        proj = our_proj

    # Parse incl/excl
    if user_incl_raw is not None and user_incl_raw != '':
        try:
            val = int(float(user_incl_raw))
            if val == 1: must_have.append(name)
            elif val == 0: must_avoid.append(name)
        except (ValueError, TypeError):
            pass

    sgm_raw = ws.cell(row=row, column=headers['SC Games']).value
    try:
        sgm = float(sgm_raw)
    except (ValueError, TypeError):
        sgm = 22.5 if team in OPENING_ROUND_CLUBS else 23.0

    bye_raw = ws.cell(row=row, column=headers['Bye']).value or '-'
    bye = int(bye_raw.replace('R', '')) if isinstance(bye_raw, str) and bye_raw.startswith('R') else 0
    games_3r = 2 if bye in (1, 2, 3, 4) else 3

    be = price / 5400
    proj_rise = int((proj - be) * PRICE_CHANGE_PER_PT * games_3r)

    r0_raw = ws.cell(row=row, column=headers['R0 Score']).value
    r0_score = int(r0_raw) if r0_raw is not None and r0_raw != '' else None

    players.append({
        'name': name, 'team': team, 'positions': positions,
        'pos_str': pos_str, 'price': int(price),
        'own': own, 'proj_avg3': proj, 'source': source,
        'sc_season_games': sgm, 'bye': bye, 'games_3r': games_3r,
        'proj_rise': proj_rise, 'r0_score': r0_score,
    })

print(f"  {len(players)} players loaded")
print(f"  MUST HAVE: {must_have}")
print(f"  MUST AVOID: {len(must_avoid)} players")
user_overrides = [p for p in players if p['source'] == 'USER']
if user_overrides:
    print(f"  User projection overrides: {len(user_overrides)}")
    for p in user_overrides[:10]:
        print(f"    {p['name']}: {p['proj_avg3']}")
    if len(user_overrides) > 10:
        print(f"    ... and {len(user_overrides) - 10} more")

# =============================================================================
# 2. OPTIMISER
# =============================================================================
def get_primary_position(positions):
    for pos in ['RUC', 'MID', 'FWD', 'DEF']:
        if pos in positions: return pos
    return 'FWD'


def optimise_team(
    players,
    salary_cap=10_000_000,
    squad_size={'DEF': 8, 'MID': 11, 'RUC': 3, 'FWD': 8, 'FLX': 1},
    keeper_thresholds={'DEF': 90, 'MID': 100, 'RUC': 100, 'FWD': 90},
    cow_rounds=8,
    w_cow_rise=3.5,
    max_per_club=None,
    must_have=None, must_avoid=None,
    proj_noise=0.0,
    label="OPTIMAL TEAM",
    quiet=False,
):
    eligible = [p for p in players if p['name'] not in (must_avoid or [])]
    if not quiet:
        print(f"\n{'='*140}")
        print(f"  {label}")
        print(f"  Eligible: {len(eligible)} | Cap: ${salary_cap:,} | Keeper: {keeper_thresholds}")
        print(f"  Cow model: {cow_rounds}rd, rise_wt={w_cow_rise}")
        print(f"{'='*140}")

    n = len(eligible)
    pos_list = ['DEF', 'MID', 'RUC', 'FWD', 'FLX']

    prob = LpProblem("SC2026", LpMaximize)
    x = [LpVariable(f"x_{i}", cat='Binary') for i in range(n)]
    y = [[LpVariable(f"y_{i}_{j}", cat='Binary') for j in range(len(pos_list))] for i in range(n)]

    obj = []
    for i, p in enumerate(eligible):
        avg = p['proj_avg3']
        if proj_noise > 0:
            avg = max(0, avg + random.gauss(0, avg * proj_noise))

        sgm = p['sc_season_games']
        rise = p['proj_rise']
        primary_pos = get_primary_position(p['positions'])
        threshold = keeper_thresholds.get(primary_pos, 90)

        if avg >= threshold:
            val = avg * sgm
        else:
            val = avg * cow_rounds + w_cow_rise * (rise / 1000)
        obj.append(val)

    prob += lpSum(x[i] * obj[i] for i in range(n))
    prob += lpSum(x[i] * eligible[i]['price'] for i in range(n)) <= salary_cap

    for j, pos in enumerate(pos_list):
        prob += lpSum(y[i][j] for i in range(n)) == squad_size.get(pos, 0)
    for i in range(n):
        prob += lpSum(y[i][j] for j in range(len(pos_list))) == x[i]
    for i, p in enumerate(eligible):
        for j, pos in enumerate(pos_list):
            if pos != 'FLX' and pos not in p['positions']:
                prob += y[i][j] == 0
    if max_per_club:
        for club in set(p['team'] for p in eligible):
            idxs = [i for i, p in enumerate(eligible) if p['team'] == club]
            prob += lpSum(x[i] for i in idxs) <= max_per_club
    if must_have:
        for name in must_have:
            idx = next((i for i, p in enumerate(eligible) if p['name'] == name), None)
            if idx is not None:
                prob += x[idx] == 1
            elif not quiet:
                print(f"  WARNING: MUST HAVE '{name}' not found in eligible pool")

    prob.solve(PULP_CBC_CMD(msg=0))
    status = LpStatus[prob.status]
    if status != 'Optimal':
        if not quiet: print(f"  STATUS: {status}")
        return status, [], 0, 0, 0, 0

    selected = []
    for i, p in enumerate(eligible):
        if x[i].varValue and x[i].varValue > 0.5:
            apos = 'UNK'
            for j, pos in enumerate(pos_list):
                if y[i][j].varValue and y[i][j].varValue > 0.5:
                    apos = pos; break
            selected.append({**p, 'assigned_pos': apos, 'obj_value': obj[i]})

    sal = sum(p['price'] for p in selected)
    rise = sum(p['proj_rise'] for p in selected)
    obj_total = sum(p['obj_value'] for p in selected)
    return status, selected, obj_total, sal, 0, rise


def print_team(selected, sal, pts, rise, obj_total, keeper_thresholds, label=""):
    pos_order = {'DEF': 0, 'MID': 1, 'RUC': 2, 'FWD': 3, 'FLX': 4}
    selected.sort(key=lambda p: (pos_order.get(p['assigned_pos'], 5), -p['price']))

    print(f"\n{'─'*155}")
    print(f"{'Pos':<5} {'Player':<26} {'NatPos':<10} {'Team':<5} {'Price':>10} {'Own%':>6} "
          f"{'Avg':>6} {'Type':<5} {'SeasonPts':>9} {'EarlyΔ':>9} {'R0':>4} {'Src':<7} {'EBye':>5}")
    print(f"{'─'*155}")

    pos_totals = defaultdict(lambda: {'n': 0, 'sal': 0, 'season_pts': 0, 'rise': 0})
    keepers, cows = [], []
    cur = None
    for p in selected:
        apos = p['assigned_pos']
        if apos != cur:
            if cur:
                t = pos_totals[cur]
                print(f"{'':>5} {'── Subtotal':<26} {'':10} {'':5} ${t['sal']:>9,} {'':>6} "
                      f"{'':>6} {'':>5} {t['season_pts']:>9.0f} ${t['rise']:>8,}")
                print()
            cur = apos

        bye_str = f"R{p['bye']}" if p['bye'] > 0 else "-"
        src = p['source'][:6]
        sgm = p['sc_season_games']
        season_pts = p['proj_avg3'] * sgm
        primary_pos = get_primary_position(p['positions'])
        threshold = keeper_thresholds.get(primary_pos, 90)
        is_keeper = p['proj_avg3'] >= threshold
        ptype = 'KEEP' if is_keeper else 'COW'
        (keepers if is_keeper else cows).append(p)
        r0 = str(int(p['r0_score'])) if p.get('r0_score') is not None else "-"

        print(f"{apos:<5} {p['name']:<26} {p['pos_str']:<10} {p['team']:<5} ${p['price']:>9,} "
              f"{p['own']:>5.1f}% {p['proj_avg3']:>6.1f} {ptype:<5} {season_pts:>9.0f} "
              f"${p['proj_rise']:>8,} {r0:>4} {src:<7} {bye_str:>5}")

        pos_totals[apos]['n'] += 1
        pos_totals[apos]['sal'] += p['price']
        pos_totals[apos]['season_pts'] += round(season_pts)
        pos_totals[apos]['rise'] += p['proj_rise']

    if cur:
        t = pos_totals[cur]
        print(f"{'':>5} {'── Subtotal':<26} {'':10} {'':5} ${t['sal']:>9,} {'':>6} "
              f"{'':>6} {'':>5} {t['season_pts']:>9.0f} ${t['rise']:>8,}")

    print(f"\n{'═'*155}")
    total_season = sum(p['proj_avg3'] * p['sc_season_games'] for p in selected)
    keeper_sal = sum(p['price'] for p in keepers)
    cow_sal = sum(p['price'] for p in cows)
    cow_rise = sum(p['proj_rise'] for p in cows)
    print(f"  TOTAL: {len(selected)} players | Salary: ${sal:,} / $10,000,000 (${10_000_000 - sal:,} remaining)")
    print(f"  Projected Season Points: {total_season:,.0f} (avg {total_season/23:.0f}/round)")
    print(f"  Structure: {len(keepers)} keepers (${keeper_sal:,}) + {len(cows)} cash cows (${cow_sal:,})")
    print(f"  Cash cow projected early rise: ${cow_rise:,}")
    print(f"  Objective Value: {obj_total:.1f}")

    high_own = [p for p in selected if p['own'] > 40]
    low_own = [p for p in selected if p['own'] < 5 and p['price'] > 200000]
    clubs = Counter(p['team'] for p in selected)
    max_club = clubs.most_common(1)[0]
    print(f"\n  Template: {len(high_own)} players >40% owned, {len(low_own)} PODs <5% owned (>$200k)")
    print(f"  Clubs: {len(clubs)} represented, max {max_club[1]} from {max_club[0]}")

    rucks = [p for p in selected if p['assigned_pos'] in ('RUC', 'FLX') and 'RUC' in p['positions']]
    if rucks:
        print(f"  Ruck structure:")
        for r in rucks:
            print(f"    {r['assigned_pos']}: {r['name']} (${r['price']:,}, avg {r['proj_avg3']:.0f})")
    print(f"{'═'*155}")


# =============================================================================
# 3. MONTE CARLO
# =============================================================================
def run_monte_carlo(players, n_sims=1000, keeper_thresholds=None, **kwargs):
    if keeper_thresholds is None:
        keeper_thresholds = {'DEF': 90, 'MID': 100, 'RUC': 100, 'FWD': 90}

    counts = Counter()
    pos_counts = defaultdict(Counter)

    sim_iter = range(n_sims)
    if tqdm is not None:
        sim_iter = tqdm(sim_iter, desc="Monte Carlo", unit="sim",
                        bar_format="{l_bar}{bar}| {n_fmt}/{total_fmt} sims [{elapsed}<{remaining}, {rate_fmt}]")

    for sim in sim_iter:
        status, team, *_ = optimise_team(
            players, keeper_thresholds=keeper_thresholds,
            proj_noise=0.15, quiet=True, **kwargs
        )
        if status == 'Optimal':
            for p in team:
                counts[p['name']] += 1
                pos_counts[p['name']][p['assigned_pos']] += 1

    results = []
    for name, count in counts.most_common():
        freq = count / n_sims
        main_pos = pos_counts[name].most_common(1)[0][0]
        pdata = next((p for p in players if p['name'] == name), None)
        results.append({
            'name': name, 'freq': freq, 'count': count,
            'main_pos': main_pos,
            'price': pdata['price'] if pdata else 0,
            'avg': pdata['proj_avg3'] if pdata else 0,
            'own': pdata['own'] if pdata else 0,
            'team': pdata['team'] if pdata else '',
        })
    return results


# =============================================================================
# 4. RUN
# =============================================================================

# Scenario 1
THRESH_1 = {'DEF': 90, 'MID': 100, 'RUC': 100, 'FWD': 90}
print("\n" + "█"*155)
print(f"█  SCENARIO 1: Keeper thresholds DEF/FWD≥90, MID/RUC≥100")
print("█"*155)
s, t1, o, sal, pts, rise = optimise_team(
    players, keeper_thresholds=THRESH_1,
    must_have=must_have, must_avoid=must_avoid,
    label="SCENARIO 1: DEF/FWD≥90, MID/RUC≥100"
)
if s == 'Optimal': print_team(t1, sal, pts, rise, o, THRESH_1)

# Scenario 2
THRESH_2 = {'DEF': 95, 'MID': 105, 'RUC': 105, 'FWD': 95}
print("\n" + "█"*155)
print(f"█  SCENARIO 2: Keeper thresholds DEF/FWD≥95, MID/RUC≥105")
print("█"*155)
s, t2, o, sal, pts, rise = optimise_team(
    players, keeper_thresholds=THRESH_2,
    must_have=must_have, must_avoid=must_avoid,
    label="SCENARIO 2: DEF/FWD≥95, MID/RUC≥105"
)
if s == 'Optimal': print_team(t2, sal, pts, rise, o, THRESH_2)

# Monte Carlo on Scenario 2
print("\n" + "█"*155)
print("█  MONTE CARLO: 1000 simulations with ±15% projection noise")
print("█"*155)

mc_results = run_monte_carlo(
    players, n_sims=1000,
    keeper_thresholds=THRESH_2,
    must_have=must_have, must_avoid=must_avoid,
)

print(f"\n{'Name':<26} {'Pos':<5} {'Team':<5} {'Price':>9} {'Avg':>5} {'Own%':>6} {'Select%':>8}")
print("-" * 80)
for r in mc_results[:50]:
    print(f"{r['name']:<26} {r['main_pos']:<5} {r['team']:<5} ${r['price']:>8,} "
          f"{r['avg']:>5.0f} {r['own']:>5.1f}% {r['freq']:>7.1%}")

# Robustness
print("\n" + "█"*155)
print("█  ROBUSTNESS: Near-misses and high-ownership omissions")
print("█"*155)

selected_names = {p['name'] for p in t1}

print(f"\nNEAR MISSES (10-80% MC selection, not in deterministic team):")
print(f"{'Name':<26} {'Pos':<5} {'Team':<5} {'Price':>9} {'Avg':>5} {'Own%':>6} {'MC%':>7}")
print("-" * 75)
for r in mc_results:
    if r['name'] not in selected_names and 0.10 <= r['freq'] <= 0.80:
        print(f"{r['name']:<26} {r['main_pos']:<5} {r['team']:<5} ${r['price']:>8,} "
              f"{r['avg']:>5.0f} {r['own']:>5.1f}% {r['freq']:>6.1%}")

print(f"\nHIGH OWNERSHIP OMISSIONS (top 50 by own%, not in team):")
all_by_own = sorted(players, key=lambda p: -p['own'])
print(f"{'Name':<26} {'Pos':<10} {'Team':<5} {'Price':>9} {'Avg':>5} {'Own%':>6} {'MC%':>7} {'Note'}")
print("-" * 100)
shown = 0
for p in all_by_own[:50]:
    if p['name'] not in selected_names and p['name'] not in must_avoid:
        mc_freq = next((r['freq'] for r in mc_results if r['name'] == p['name']), 0)
        if mc_freq > 0.5: note = "STRONG near-miss"
        elif mc_freq > 0.1: note = "Marginal"
        elif mc_freq > 0: note = "Rare selection"
        else: note = "Never selected"
        print(f"{p['name']:<26} {p['pos_str']:<10} {p['team']:<5} ${p['price']:>8,} "
              f"{p['proj_avg3']:>5.0f} {p['own']:>5.1f}% {mc_freq:>6.1%} {note}")
        shown += 1
        if shown >= 25: break
